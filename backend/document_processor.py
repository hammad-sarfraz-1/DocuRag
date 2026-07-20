import io
import uuid
from datetime import datetime, timezone
from typing import List

from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from langchain_experimental.text_splitter import SemanticChunker
from langchain_text_splitters import RecursiveCharacterTextSplitter

from backend.config import Config
from backend.embeddings import LANGCHAIN_EMBEDDING_FN

# Same embedding model used for retrieval/cache, so chunk boundaries are cut
# where meaning actually shifts rather than at a fixed character count.
_semantic_chunker = SemanticChunker(LANGCHAIN_EMBEDDING_FN)

# SemanticChunker has no size ceiling — pathological input (e.g. long runs of
# near-duplicate paragraphs) can land in one giant chunk. Fallback splitter
# only for chunks that exceed MAX_CHUNK_CHARS, with overlap so context isn't
# lost at the new boundary.
_fallback_splitter = RecursiveCharacterTextSplitter(
    chunk_size=Config.MAX_CHUNK_CHARS, chunk_overlap=200
)


# ---------------------------------------------------------------------------
# OCR (optional — disabled if EasyOCR not available or config says so)
# ---------------------------------------------------------------------------
_ocr_reader = None


def _init_ocr():
    global _ocr_reader
    if _ocr_reader is None and Config.USE_OCR:
        try:
            import easyocr

            _ocr_reader = easyocr.Reader(Config.OCR_LANG, gpu=False)
        except ImportError:
            pass  # OCR not available
    return _ocr_reader


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text from a PDF. Falls back to OCR when standard extraction
    yields little or no text."""
    reader = PdfReader(io.BytesIO(file_bytes))
    text = "\n".join([page.extract_text() or "" for page in reader.pages])
    text = text.strip()

    if Config.USE_OCR and (not text or len(text) < 50):
        ocr = _init_ocr()
        if ocr is not None:
            try:
                from pdf2image import convert_from_bytes
                import numpy as np

                images = convert_from_bytes(file_bytes)
                ocr_lines = []
                for img in images:
                    img_np = np.array(img)
                    result = ocr.readtext(img_np, detail=0)
                    ocr_lines.extend(result)
                text = "\n".join(ocr_lines)
            except ImportError:
                pass  # pdf2image or numpy not installed
    return text


def extract_text_from_docx(file_bytes: bytes) -> str:
    doc = Document(io.BytesIO(file_bytes))
    return "\n".join([para.text for para in doc.paragraphs])


def extract_text_from_txt(file_bytes: bytes) -> str:
    return file_bytes.decode("utf-8")


def extract_text_from_excel(file_bytes: bytes) -> str:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append("\t".join(cells))
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".xlsx"}


def extract_text(file_bytes: bytes, filename: str) -> str:
    """Route a file to the correct parser based on its extension."""
    if filename.lower().endswith(".pdf"):
        return extract_text_from_pdf(file_bytes)
    elif filename.lower().endswith(".docx"):
        return extract_text_from_docx(file_bytes)
    elif filename.lower().endswith(".txt"):
        return extract_text_from_txt(file_bytes)
    elif filename.lower().endswith(".xlsx"):
        return extract_text_from_excel(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {filename}")


_CHUNK_OVERLAP_CHARS = 50


def chunk_text(text: str) -> List[str]:
    """Split text into chunks at semantic boundaries (where consecutive
    sentences' embeddings diverge), instead of a fixed character count.

    Any resulting chunk longer than Config.MAX_CHUNK_CHARS is further split
    by RecursiveCharacterTextSplitter, since SemanticChunker alone has no
    size ceiling and can group pathological (e.g. repetitive) text into one
    oversized chunk.

    Each chunk after the first is then prefixed with the last
    _CHUNK_OVERLAP_CHARS of the chunk before it, so a query whose answer
    spans a chunk boundary still finds it in both chunks' embeddings.
    """
    chunks = _semantic_chunker.split_text(text)
    result = []
    for c in chunks:
        if not c.strip():
            continue
        if len(c) > Config.MAX_CHUNK_CHARS:
            result.extend(_fallback_splitter.split_text(c))
        else:
            result.append(c)
    result = [c for c in result if c.strip()]

    for i in range(1, len(result)):
        overlap = result[i - 1][-_CHUNK_OVERLAP_CHARS:]
        result[i] = overlap + result[i]
    return result


def build_chunk_metadata(chunks: List[str], source: str) -> List[dict]:
    """Build per-chunk metadata dicts (used for source citation).

    document_id/upload_date are stamped fresh on every upload, so a
    re-upload of the same filename gets a new id and the latest timestamp —
    add_documents() replaces the old chunks by `source`, so stale chunks
    from a prior version never linger alongside the new ones.
    """
    document_id = str(uuid.uuid4())
    upload_date = datetime.now(timezone.utc).isoformat()
    return [
        {
            "source": source,
            "chunk_index": i,
            "total_chunks": len(chunks),
            "document_id": document_id,
            "upload_date": upload_date,
        }
        for i in range(len(chunks))
    ]
